import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
import requests
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

class DefectDojoGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("DefectDojo Data Fetcher")
        # Set medium-sized window
        self.root.geometry("400x250")
        self.root.resizable(False, False)
        
        # Define larger font
        label_font = ("Arial", 12)
        button_font = ("Arial", 12)
        
        # Set default dates (today is June 16, 2025, Monday)
        today = datetime.now()
        is_monday = today.weekday() == 0
        default_end = today - timedelta(days=1)  # Yesterday: 2025-06-15
        default_start = today - timedelta(days=3) if is_monday else default_end  # Friday: 2025-06-13 if Monday
        
        # Create GUI elements with increased padding and size
        tk.Label(root, text="Start Date:", font=label_font).grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.start_date = DateEntry(root, date_pattern='yyyy-mm-dd', 
                                  year=default_start.year, 
                                  month=default_start.month, 
                                  day=default_start.day,
                                  font=label_font, width=15)
        self.start_date.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(root, text="End Date:", font=label_font).grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.end_date = DateEntry(root, date_pattern='yyyy-mm-dd',
                                 year=default_end.year,
                                 month=default_end.month,
                                 day=default_end.day,
                                 font=label_font, width=15)
        self.end_date.grid(row=1, column=1, padx=10, pady=10)

        self.start_button = ttk.Button(root, text="Start", command=self.fetch_data, style="TButton")
        self.start_button.grid(row=2, column=0, columnspan=2, pady=20)

        self.status_label = tk.Label(root, text="", font=label_font, wraplength=350)
        self.status_label.grid(row=3, column=0, columnspan=2, pady=10)

        # Configure button style
        style = ttk.Style()
        style.configure("TButton", font=button_font)

    def fetch_data(self):
        self.status_label.config(text="Fetching data...")
        self.start_button.config(state='disabled')
        
        try:
            # API configuration
            url = "https://demo.defectdojo.org/api/v2/engagements/"
            headers = {
                "Authorization": "Token DOJOTOKEN"  # Replace with actual token
            }
            
            # Get date range
            start_date = self.start_date.get_date()
            end_date = self.end_date.get_date()
            start_date_str = start_date.strftime('%Y-%m-%d')
            end_date_str = end_date.strftime('%Y-%m-%d')
            
            # Fetch data
            params = {
                'status': 'Completed',
                'tags': 'pci',
                'updated__gte': start_date_str,
                'updated__lte': end_date_str
            }
            
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()['results']
            
            # Process data for Patches file
            patches_data = []
            for item in data:
                if 'pci' in item.get('tags', []) and item.get('status') == 'Completed':
                    # Set 'Approved' for blank/null commit_hash or any value other than 'Approved with Exception' or 'Rejected'
                    commit_hash = item.get('commit_hash')
                    comment = 'Approved' if not commit_hash or commit_hash not in ['Approved with Exception', 'Rejected'] else commit_hash
                    patches_data.append({
                        'Patch': item.get('name', ''),
                        'IR': item.get('version', ''),
                        'Created': item.get('created', '').split('T')[0] if item.get('created') else '',
                        'Completed': item.get('updated', '').split('T')[0] if item.get('updated') else '',
                        'Comment': comment
                    })
            
            # Create Patches Excel file
            patches_df = pd.DataFrame(patches_data)
            patches_filename = f"Patches_Date_{start_date_str}_{end_date_str}.xlsx"
            patches_df.to_excel(patches_filename, index=False)
            
            # Load existing final.xlsx if it exists
            final_excel_path = 'final.xlsx'
            if os.path.exists(final_excel_path):
                existing_df = pd.read_excel(final_excel_path)
                # Ensure 'Patches' column exists and has the correct categories
                expected_patches = ['Approved', 'Approved with Exception', 'Rejected', 'Total']
                if not existing_df.empty and 'Patches' in existing_df.columns:
                    existing_patches = existing_df['Patches'].tolist()
                    if existing_patches != expected_patches:
                        # Reinitialize if Patches column is incorrect
                        existing_df = pd.DataFrame({
                            'Patches': expected_patches
                        })
                else:
                    existing_df = pd.DataFrame({
                        'Patches': expected_patches
                    })
            else:
                # Initialize with Patches column including Total
                existing_df = pd.DataFrame({
                    'Patches': ['Approved', 'Approved with Exception', 'Rejected', 'Total']
                })
            
            # Create or update data for the selected date range
            new_data = {}
            current_date = start_date
            while current_date <= end_date:
                date_str = current_date.strftime('%d-%b-%Y')
                daily_patches = patches_df[patches_df['Completed'] == current_date.strftime('%Y-%m-%d')]
                commit_counts = daily_patches['Comment'].value_counts().to_dict()
                if 'Approved' not in commit_counts:
                    commit_counts['Approved'] = 0
                if 'Approved with Exception' not in commit_counts:
                    commit_counts['Approved with Exception'] = 0
                if 'Rejected' not in commit_counts:
                    commit_counts['Rejected'] = 0
                total = sum([commit_counts.get('Approved', 0),
                             commit_counts.get('Approved with Exception', 0),
                             commit_counts.get('Rejected', 0)])
                new_data[date_str] = [
                    commit_counts.get('Approved', 0),
                    commit_counts.get('Approved with Exception', 0),
                    commit_counts.get('Rejected', 0),
                    total
                ]
                current_date += timedelta(days=1)
            
            # Update or append date columns in existing_df
            for date_str, counts in new_data.items():
                existing_df[date_str] = counts
            
            # Save to Excel with formatting
            existing_df.to_excel(final_excel_path, index=False)
            
            # Apply formatting using openpyxl
            wb = load_workbook(final_excel_path)
            ws = wb.active
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to all cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    # Center-align date columns (all columns except 'Patches')
                    if cell.column > 1:
                        cell.alignment = Alignment(horizontal='center')
                        # Hide zeros for non-Total rows
                        if cell.row < ws.max_row and cell.value == 0:
                            cell.value = None
                    # Bold the Total row (including 'Total' label)
                    if cell.row == ws.max_row:
                        cell.font = Font(bold=True)
                    # Apply border to all cells
                    cell.border = thin_border
            
            wb.save(final_excel_path)
            
            self.status_label.config(text=f"Success! Created {patches_filename} and updated final.xlsx")
            
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}")
        
        finally:
            self.start_button.config(state='normal')

if __name__ == "__main__":
    root = tk.Tk()
    app = DefectDojoGUI(root)
    root.mainloop()